import os
from flask import Flask, request, render_template, send_file
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import re

app = Flask(__name__)

# Define directories for uploads and reports
UPLOAD_FOLDER = 'uploads'
REPORTS_FOLDER = 'reports'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(REPORTS_FOLDER, exist_ok=True)

@app.route('/', methods=['GET', 'POST'])
def upload_files():
    if request.method == 'POST':
        # Get uploaded files
        pending_users_file = request.files.get('pending_users')
        course_status_files = request.files.getlist('course_status')

        if not pending_users_file or not course_status_files:
            return "Please upload both Pending Users and Course Status files!", 400

        # Save uploaded files
        pending_users_path = os.path.join(UPLOAD_FOLDER, pending_users_file.filename)
        pending_users_file.save(pending_users_path)

        course_status_paths = []
        for file in course_status_files:
            file_path = os.path.join(UPLOAD_FOLDER, file.filename)
            file.save(file_path)
            course_status_paths.append(file_path)

        # Process files
        report_path = process_files(pending_users_path, course_status_paths)

        # Send the generated report back to the user
        return send_file(report_path, as_attachment=True)

    return render_template('upload.html')

def process_files(pending_users_path, course_status_paths):
    # Load the pending users dataset
    pending_users_df = pd.read_csv(pending_users_path)
    pending_users_df['Email'] = pending_users_df['Email'].str.strip().str.lower()
    pending_users_df['Email Domain'] = pending_users_df['Email'].str.split('@').str[1]

    # Template path (ensure this exists)
    template_path = 'template2.xlsx'

    combined_reports = []

    for course_status_path in course_status_paths:
        new_course_status_df = pd.read_csv(course_status_path)
        new_course_status_df['Email'] = new_course_status_df['Email'].str.strip().str.lower()
        new_course_status_df['Email Domain'] = new_course_status_df['Email'].str.split('@').str[1]

        # Collect all unique emails from the course status dataset
        course_emails_set = set(new_course_status_df['Email'].dropna().unique())

        # Filter pending users
        filtered_pending_users = pending_users_df[~pending_users_df['Email'].isin(course_emails_set)]
        pending_users = filtered_pending_users[['Email', 'Last invite sent at']].copy()
        pending_users['Full name'] = 'Pending User'
        pending_users['Course name'] = 'Pending Course'
        pending_users['Progress'] = 'Not started'
        pending_users['Score'] = '0%'
        pending_users['Course completion %'] = '0%'
        pending_users['Enrolled'] = '-'
        pending_users['Started'] = '-'
        pending_users['Last accessed'] = '-'
        pending_users['Completed'] = 'Invite last sent: ' + filtered_pending_users['Last invite sent at']

        # Combine course status and pending users
        combined_df = pd.concat([new_course_status_df, pending_users], ignore_index=True)

        # Debug: Ensure Progress column exists
        print("Combined DataFrame Columns:", combined_df.columns.tolist())

        if 'Progress' not in combined_df.columns:
            raise KeyError("The 'Progress' column is missing in the combined data!")

        # Define the 'Status' column based on 'Progress'
        def determine_status(row):
            if row['Progress'] == 'Passed':
                return 'Passed'
            elif row['Progress'] == 'In Progress':
                return 'In Progress'
            else:
                return 'Not started'

        combined_df['Status'] = combined_df.apply(determine_status, axis=1)

        # Define the 'Completed' column
        def determine_completed(row):
            if row['Status'] == 'Passed':
                return f"Completed: {row['Completed'].split(' ')[0]}" if pd.notna(row['Completed']) and row['Completed'] != '-' else 'Completed: Unknown'
            elif row['Status'] == 'In Progress':
                return f"Last accessed course: {row['Last accessed'].split(' ')[0]}" if pd.notna(row['Last accessed']) and row['Last accessed'] != '-' else 'Last accessed course: Unknown'
            else:
                return row['Completed']

        combined_df['Completed'] = combined_df.apply(determine_completed, axis=1)

        # Keep only the required columns
        final_df = combined_df[['Full name', 'Email', 'Course name', 'Status', 'Completed', 'Score']].copy()

        # Add a percentage sign to the 'Score' column
        final_df['Score'] = final_df['Score'].str.rstrip('%').astype(float).round(0).astype(int).astype(str) + '%'

        # Rename columns
        final_df.rename(columns={'Full name': 'User'}, inplace=True)

        # Save the report
        wb = load_workbook(template_path)
        ws = wb.active
        start_row = 13
        start_col = 2
        for r_idx, row in final_df.iterrows():
            for c_idx, value in enumerate(row):
                ws.cell(row=start_row + r_idx, column=start_col + c_idx, value=value)

        today_date = datetime.now().strftime('%Y-%m-%d')
        course_name = re.sub(r'[<>:"/\\|?*]', '_', combined_df['Course name'].iloc[0])
        output_file = os.path.join(REPORTS_FOLDER, f"Everbright_Report_{course_name}_{today_date}.xlsx")
        wb.save(output_file)

        combined_reports.append(output_file)

    return combined_reports[0]  # Return the first report

if __name__ == '__main__':
    app.run(debug=True)
