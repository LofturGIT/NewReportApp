import os
import pandas as pd
from flask import Flask, render_template, request, send_file, session, redirect, url_for
from werkzeug.utils import secure_filename
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

app = Flask(__name__)
app.secret_key = 'supersecretkey'

UPLOAD_FOLDER = 'uploads'
TEMPLATE_PATH = 'template2.xlsx'
LOGO_PATH = 'logo.png'
REPORTS_FOLDER = 'reports'

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(REPORTS_FOLDER, exist_ok=True)

PENDING_USER_FILENAME = 'pending_users.csv'

def get_pending_users_path():
    return os.path.join(UPLOAD_FOLDER, PENDING_USER_FILENAME)

def get_status(row):
    completed = str(row.get('Completed', '')).strip()
    last_accessed = str(row.get('Last accessed', '')).strip()

    if completed and completed != '-':
        return 'Passed'
    elif last_accessed and last_accessed != '-':
        return 'In Progress'
    else:
        return 'Not Started'

def get_completed_date(row):
    status = row['Status']
    try:
        if status == 'Passed':
            date = pd.to_datetime(row['Completed'], errors='coerce', dayfirst=True)
        elif status == 'In Progress':
            date = pd.to_datetime(row['Last accessed'], errors='coerce', dayfirst=True)
        elif status == 'Not Started':
            date = pd.to_datetime(row['Enrolled'], errors='coerce', dayfirst=True)
        else:
            return ''
        if pd.notna(date):
            return date.strftime('%d/%m/%Y')
    except:
        return ''
    return ''

def process_files(pending_users_path, course_status_path, selected_domains, selected_group):
    # Load pending users
    pending_users_df = pd.read_csv(pending_users_path)
    pending_users_df['Email Domain'] = pending_users_df['Email'].str.split('@').str[1]
    filtered_pending = pending_users_df[pending_users_df['Email Domain'].isin(selected_domains)].copy()

    # Standardize emails for matching
    filtered_pending['Email'] = filtered_pending['Email'].str.lower().str.strip()
    pending_emails_set = set(filtered_pending['Email'])

    # Load report
    df = pd.read_csv(course_status_path)
    df.columns = df.columns.str.strip()
    df['Email'] = df['Email'].str.lower().str.strip()

    # Identify users who haven't completed
    completed_emails = set(df['Email'].dropna().unique())
    filtered_pending = filtered_pending[~filtered_pending['Email'].isin(completed_emails)].copy()

    # Format pending user fields
    filtered_pending['User'] = 'Pending user'
    filtered_pending['Course name'] = 'Pending course'
    filtered_pending['Status'] = 'Not Started'
    filtered_pending['Score'] = '0%'

    # ✅ FIX: Use dayfirst=True for proper British format
    filtered_pending['Date'] = filtered_pending['Last invite sent at'].apply(
        lambda x: pd.to_datetime(x, errors='coerce', dayfirst=True).strftime('%d/%m/%Y')
        if pd.notna(pd.to_datetime(x, errors='coerce', dayfirst=True)) else ''
    )

    pending_final = filtered_pending[['User', 'Email', 'Course name', 'Status', 'Date', 'Score']]

    # Prepare main dataframe
    df['User'] = df['Full name']
    df['Status'] = df.apply(get_status, axis=1)
    df['Date'] = df.apply(get_completed_date, axis=1)
    df['Score'] = df['Score'].str.rstrip('%').astype(float).round(0).astype(int).astype(str) + '%'

    completed_final = df[['User', 'Email', 'Course name', 'Status', 'Date', 'Score']]

    combined = pd.concat([completed_final, pending_final], ignore_index=True)

    # Load template and write to Excel
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    start_row = 13
    for idx, col in enumerate(['User', 'Email', 'Course name', 'Status', 'Date', 'Score']):
        for i, value in enumerate(combined[col], start=start_row):
            ws.cell(row=i, column=idx + 2, value=value)

    # Add logo if it exists
    if os.path.exists(LOGO_PATH):
        img = Image(LOGO_PATH)
        img.anchor = "D4"
        ws.add_image(img)

    # Save report
    course_name = df['Course name'].iloc[0].replace(" ", "_").replace("/", "-")
    group_part = selected_group.replace(" ", "_").replace("/", "-")
    date_str = datetime.now().strftime("%d-%m-%Y")
    output_path = os.path.join(REPORTS_FOLDER, f"Report_{course_name}_{group_part}_{date_str}.xlsx")
    wb.save(output_path)

    return output_path

@app.route('/', methods=['GET', 'POST'])
def upload_pending():
    if request.method == 'POST':
        file = request.files['pending_users']
        if file:
            pending_path = get_pending_users_path()
            file.save(pending_path)
            return redirect(url_for('upload_report'))
    return render_template('upload_pending.html')

@app.route('/upload_report', methods=['GET', 'POST'])
def upload_report():
    if not os.path.exists(get_pending_users_path()):
        return redirect(url_for('upload_pending'))

    if request.method == 'POST':
        report_file = request.files['report_file']
        if not report_file:
            return "Please upload a report file."

        path = os.path.join(UPLOAD_FOLDER, secure_filename(report_file.filename))
        report_file.save(path)

        df = pd.read_csv(path)
        domains = df['Email'].str.split('@').str[1].dropna().unique().tolist()

        try:
            groups_series = df['User’s groups'].dropna().str.split(',')
            all_groups = set(group.strip() for sublist in groups_series for group in sublist)
            groups = sorted(all_groups)
        except KeyError:
            groups = []

        return render_template('select_domains_and_group.html', report_path=path, domains=domains, groups=groups)

    return render_template('upload_report.html')

@app.route('/generate_report', methods=['POST'])
def generate_report():
    report_path = request.form['report_path']
    selected_domains = request.form.getlist('selected_domains')
    selected_group = request.form.get('selected_group')

    if not selected_domains or not selected_group:
        return "You must select at least one domain and one group."

    output = process_files(get_pending_users_path(), report_path, selected_domains, selected_group)
    return send_file(output, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
